from aiogram import types
import os
import re
import csv
from openpyxl import load_workbook
from datetime import datetime, timedelta
from aiogram.types import ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton, ContentType
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.dispatcher.filters import Command

from database import check_tracking_status, get_tracking_status, get_tracking_arriveDate, get_track_link
from database import link_track_toPhone, myOrderList, get_admin_code, save_to_db, export_user_data, update_tracking_status, notify_users_about_arrivals, delete_from_db

class Form(StatesGroup):
    check_tracking = State()
    adminPage = State()
    updateList = State()
    changeStatus = State()
    deleteList = State()
    admin = State()
    start = State()

# Start command: Check if user is already logged in or ask for credentials
async def start(update: types.Message):
    keyboard = [
        ["📝 Мои трек-коды", "💵 Цены"], 
        ["🚫 Запрещённые товары", "💬 Связаться с нами"],
        ["📍 Получить адрес", "👤 Подписаться"],
        ["🎓 Бесплатное обучение", "💱 Курс обмена (RMB/TJS)"],
        ["📦 Отслеживать трек-коды"]
        ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.answer(f"👋 Привет! Добро пожаловать в Cargo Bot!\nЗдесь вы можете легко проверить статус вашего груза. Просто отправьте трек-номер и получите информацию! 📦🔍", reply_markup=reply_markup)
    return

# When user presses "📦 Отслеживать доставку" button, hide buttons and ask for tracking code
async def check_trackCodeMessage(update: types.Message):
    keyboard = [["Отмена/Выход"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.answer("📦 Введите ваш код отслеживания:", reply_markup=reply_markup)
    await Form.check_tracking.set()

# Check track code when the user sends a tracking code
async def check_trackCode(update: types.Message, state: FSMContext):
    track_code = update.text.strip()
    uid = update.from_user.id  # Get user ID
    
    if track_code == "Отмена/Выход":
        keyboard = [
            ["📝 Мои трек-коды", "💵 Цены"], 
            ["🚫 Запрещённые товары", "💬 Связаться с нами"],
            ["📍 Получить адрес", "👤 Подписаться"],
            ["🎓 Бесплатное обучение", "💱 Курс обмена (RMB/TJS)"],
            ["📦 Отслеживать трек-коды"]
            ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
        await state.finish()
    else:
        if not await check_tracking_status(track_code):
            await update.answer("❌ Код отслеживания не найден в базе данных. Попробуйте еще раз.", reply_markup=ReplyKeyboardRemove())
            await Form.check_tracking.set()  
            return

        # Store track_code and uid in state
        await state.update_data(track_code=track_code, uid=uid)

        # Get additional tracking info
        track_status = await get_tracking_status(track_code)
        track_arriveDate = await get_tracking_arriveDate(track_code)

        date_format = "%d.%m.%Y"
        date_object = datetime.strptime(track_arriveDate, date_format)
        new_date = date_object + timedelta(days=25)

        track_arriveDateTJ = new_date.strftime("%d.%m.%Y")

        if track_status == 'active': track_status = 'Прибыло на наш склад в Китае и отправлено на склад в Таджикистане.'
        if track_status == 'arrived': track_status = 'Товар прибыл на наш склад в Таджикистане. Если мы вам все еще не позвонили, свяжитесь с администратором.'

        keyboard = [
            ["📝 Мои трек-коды", "💵 Цены"], 
            ["🚫 Запрещённые товары", "💬 Связаться с нами"],
            ["📍 Получить адрес", "👤 Подписаться"],
            ["🎓 Бесплатное обучение", "💱 Курс обмена (RMB/TJS)"],
            ["📦 Отслеживать трек-коды"]
            ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.answer(f"✅ Статус трек-кода: {track_status}\n📅 Дата прибытия на склад в Китае: {track_arriveDate}\n📅 Предполагаемая дата прибытия на наш склад в Таджикистане: ~ {track_arriveDateTJ}", reply_markup=reply_markup)
        if await get_track_link(track_code) == "unlinked":
            keyboard = InlineKeyboardMarkup().add(
                InlineKeyboardButton("✅ Да", callback_data="link_track_yes"),
                InlineKeyboardButton("❌ Нет", callback_data="link_track_no")
            )
            await update.answer("Хотите привязать этот трек к своему телефону?", reply_markup=keyboard)

        await state.set_state(None)  # Instead of await state.finish()

# Callback handler for "Yes" button
async def link_track_yes(callback_query: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()

    track_code = data.get("track_code")
    uid = data.get("uid")

    await link_track_toPhone(track_code, uid)
    await callback_query.message.edit_text("✅ Код отслеживания привязан к вашему телефону!")

# Callback handler for "No" button
async def link_track_no(callback_query: types.CallbackQuery):
    await callback_query.message.edit_text("❌ Код отслеживания не был привязан.")

# My Orders button
async def myOrders(update: types.Message):
    uid = update.from_user.id

    myList = await myOrderList(uid)  # Await the result of the async function

    if myList:
        tracking_codes = "\n".join(f"`{row[0]}`" for row in myList)  # Format each line in monospace
        message_text = f"📦 *Список заказов (привязан к этому устройству):*\n\n{tracking_codes}"
    else:
        message_text = "Заказы не найдены."

    await update.answer(
    message_text,
    parse_mode='markdown'
)

async def contactUs(update: types.Message, state: FSMContext):
    await update.answer("🔹Инстагарам: rahmonov.me✅\n🔹Админ: @ismoil_rahmonov")
    await state.finish()




# Admin Login
async def admin(update: types.Message):

    # Hide buttons and ask for tracking code
    await update.answer("Введите секретный код:", reply_markup=ReplyKeyboardRemove())
    await Form.adminPage.set()

# Admin Page
async def adminPage(update: types.Message, state: FSMContext):
    secret_code = update.text.strip()

    if secret_code == "/start": await Form.start.set()
    else:
        if not await get_admin_code(secret_code):
            await update.answer("❌ Неверный секретный код. Попробуйте снова...")
            return
        
        keyboard = [
                ["📦 Обновить список трек-кодов", "📈 Статистика"],
                ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
                ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
                ["📤 Экспорт данных"]
                ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.answer("🎉 Добро пожаловать в панель администратора!")
        await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
        await state.finish()

# Обновить список трек-кодов Message
async def updatePage(update: types.Message, state: FSMContext):
    keyboard = [["Отмена/Выход"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.answer("Пожалуйста, отправьте Excel-файл новых трек-кодов:", reply_markup=reply_markup)
    await Form.updateList.set()

async def cancel_task(update: types.Message, state: FSMContext):
    keyboard = [
            ["📦 Обновить список трек-кодов", "📈 Статистика"],
            ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
            ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
            ["📤 Экспорт данных"]
            ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
    await state.finish()

async def updateList(update: types.Message, state: FSMContext):
    from bot import bot  # Import bot inside function to avoid circular import

    # ✅ If user sends unsupported content (photo, video, location, etc.), ignore it
    if update.content_type not in [ContentType.DOCUMENT, ContentType.TEXT]:
        print(f"❌ Получен недействительный контент: {update.content_type}")
        await update.reply("⚠️ Я принимаю только файлы Excel. Пожалуйста, отправьте действительный файл.")
        return

    # ✅ If user sends text instead of file
    if update.content_type == ContentType.TEXT:
        user_text = update.text.lower()
        print(f"📝 User sent text: {user_text}")

        if user_text == "отмена/выход":
            keyboard = [
            ["📦 Обновить список трек-кодов", "📈 Статистика"],
            ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
            ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
            ["📤 Экспорт данных"]
            ]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
            await state.finish()
        else:
            await update.reply("⚠️ Я принимаю только файлы Excel. Пожалуйста, отправьте действительный файл.")
        return  

    # ✅ If user sends an Excel document
    document = update.document
    file_name = document.file_name.lower()
    allowed_extensions = [".xls", ".xlsx", ".csv"]

    if not any(file_name.endswith(ext) for ext in allowed_extensions):
        print(f"❌ Неверный тип файла: {file_name}")
        await update.reply("❌ Я принимаю только файлы Excel (.xls, .xlsx, .csv). Пожалуйста, отправьте правильный файл!")
        return  

    print(f"📂 Valid Excel file received: {file_name}")

    # ✅ Download the file
    file_id = document.file_id
    file = await bot.get_file(file_id)

    folder_path = "temp_cargo/"
    os.makedirs(folder_path, exist_ok=True)  # Ensure folder exists
    file_path = os.path.join(folder_path, file_name)

    await bot.download_file(file.file_path, file_path)
    print(f"✅ File downloaded: {file_path}")

    # ✅ Process the Excel file
    try:
        extracted_data = []
        
        if file_name.endswith(".xlsx"):
            wb = load_workbook(filename=file_path, data_only=True)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]  # Read first row as headers
            if "Трек-коды" not in headers or "Код номер" not in headers:
                await update.reply("❌ Неверный формат Excel! В файле должны быть столбцы ‘Трек-коды’ и ‘Код номер’.")
                return
            
            track_col = headers.index("Трек-коды") + 1
            code_col = headers.index("Код номер") + 1
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                tracking_code = str(row[track_col - 1]).strip() if row[track_col - 1] else ""
                cargo_code = str(row[code_col - 1]).strip() if row[code_col - 1] else ""
                
                if tracking_code and cargo_code and "nan" not in tracking_code.lower():
                    match = re.match(r"^TJ-(\d+)-(\d+)$", cargo_code)
                    if match:
                        day = match.group(1)
                        arrival_date = f"{int(day):02d}.{datetime.now().month:02d}.{datetime.now().year}"
                        extracted_data.append((tracking_code, arrival_date))
            
        elif file_name.endswith(".csv"):
            with open(file_path, encoding='utf-8') as f:
                reader = csv.DictReader(f)
                if "Трек-коды" not in reader.fieldnames or "Код номер" not in reader.fieldnames:
                    await update.reply("❌ Неверный формат CSV! В файле должны быть столбцы ‘Трек-коды’ и ‘Код номер’.")
                    return
                
                for row in reader:
                    tracking_code = row.get("Трек-коды", "").strip()
                    cargo_code = row.get("Код номер", "").strip()
                    
                    if tracking_code and cargo_code and "nan" not in tracking_code.lower():
                        match = re.match(r"^TJ-(\d+)-(\d+)$", cargo_code)
                        if match:
                            day = match.group(1)
                            arrival_date = f"{int(day):02d}.{datetime.now().month:02d}.{datetime.now().year}"
                            extracted_data.append((tracking_code, arrival_date))
        
        # ✅ Save only unique data
        if extracted_data:
            await save_to_db(extracted_data)
            unique_count = len(extracted_data)
            await update.reply(f"✅ {unique_count} новые записи сохранены. Дубликаты пропущены!")
            keyboard = [
                ["📦 Обновить список трек-кодов", "📈 Статистика"],
                ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
                ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
                ["📤 Экспорт данных"]
            ]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
            await state.finish()
        else:
            await update.reply("⚠️ В файле не найдено новых действительных кодов треков.")
            keyboard = [
            ["📦 Обновить список трек-кодов", "📈 Статистика"],
            ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
            ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
            ["📤 Экспорт данных"]
            ]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
            await state.finish()

        # ✅ Delete file after processing
        os.remove(file_path)
    
    except Exception as e:
        print(f"❌ Error processing Excel file: {e}")
        await update.reply("❌ Не удалось обработать файл. Проверьте формат и попробуйте еще раз.")

async def exportUserData(update: types.Message):
    """ Command to trigger user data export. """
    await export_user_data(update.bot, update.chat.id)

async def changeStatus(update: types.Message):
    keyboard = [["Отмена/Выход"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.answer("Пожалуйста, отправьте дату трек-кодов (полученных в Китае), которые прибыли в Таджикистан:\nНапример: 01.03.2025 - 01.04.2025", reply_markup=reply_markup)
    await Form.changeStatus.set()

async def changeStatusTrack(update: types.Message, state: FSMContext):
    date = update.text.strip()

    if date == "Отмена/Выход":
        keyboard = [
        ["📦 Обновить список трек-кодов", "📈 Статистика"],
        ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
        ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
        ["📤 Экспорт данных"]
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
        await state.finish()
    else:
        result = await update_tracking_status(date)
        await update.reply(result)
        keyboard = [
        ["📦 Обновить список трек-кодов", "📈 Статистика"],
        ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
        ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
        ["📤 Экспорт данных"]
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
        await state.finish()

async def notify_users(update: types.Message, state: FSMContext):
    """ Trigger the notification process for arrived track codes """
    
    # ✅ Pass `update.bot` to notify_users_about_arrivals
    result = await notify_users_about_arrivals(update.bot)
    
    await update.answer(result)

    # ✅ Show keyboard after notifications are sent
    keyboard = [
        ["📦 Обновить список трек-кодов", "📈 Статистика"],
        ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
        ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
        ["📤 Экспорт данных"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    
    await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
    await state.finish()

async def deleteListMessage(update: types.Message, state: FSMContext):
    keyboard = [["Отмена/Выход"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.answer("Пожалуйста, отправьте Excel-файл с готовыми трек-кодами:", reply_markup=reply_markup)
    await Form.deleteList.set()

async def deleteList(update: types.Message, state: FSMContext):
    from bot import bot  # Import bot inside function to avoid circular import

    # ✅ If user sends unsupported content (photo, video, location, etc.), ignore it
    if update.content_type not in [ContentType.DOCUMENT, ContentType.TEXT]:
        print(f"❌ Получен недействительный контент: {update.content_type}")
        await update.reply("⚠️ Я принимаю только файлы Excel. Пожалуйста, отправьте действительный файл.")
        return

    # ✅ If user sends text instead of file
    if update.content_type == ContentType.TEXT:
        user_text = update.text.lower()
        print(f"📝 User sent text: {user_text}")

        if user_text == "отмена/выход":
            keyboard = [
            ["📦 Обновить список трек-кодов", "📈 Статистика"],
            ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
            ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
            ["📤 Экспорт данных"]
            ]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
            await state.finish()
        else:
            await update.reply("⚠️ Я принимаю только файлы Excel. Пожалуйста, отправьте действительный файл.")
        return  

    # ✅ If user sends an Excel document
    document = update.document
    file_name = document.file_name.lower()
    allowed_extensions = [".xls", ".xlsx", ".csv"]

    if not any(file_name.endswith(ext) for ext in allowed_extensions):
        print(f"❌ Неверный тип файла: {file_name}")
        await update.reply("❌ Я принимаю только файлы Excel (.xls, .xlsx, .csv). Пожалуйста, отправьте правильный файл!")
        return  

    print(f"📂 Valid Excel file received: {file_name}")

    # ✅ Download the file
    file_id = document.file_id
    file = await bot.get_file(file_id)

    folder_path = "temp_cargo/"
    os.makedirs(folder_path, exist_ok=True)  # Ensure folder exists
    file_path = os.path.join(folder_path, file_name)

    await bot.download_file(file.file_path, file_path)
    print(f"✅ File downloaded: {file_path}")

    # ✅ Process the Excel file
    try:
        extracted_data = []
        
        if file_name.endswith(".xlsx"):
            wb = load_workbook(filename=file_path, data_only=True)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]  # Read first row as headers
            if "Трек-коды" not in headers or "Код номер" not in headers:
                await update.reply("❌ Неверный формат Excel! В файле должны быть столбцы ‘Трек-коды’ и ‘Код номер’.")
                return
            
            track_col = headers.index("Трек-коды") + 1
            code_col = headers.index("Код номер") + 1
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                tracking_code = str(row[track_col - 1]).strip() if row[track_col - 1] else ""
                cargo_code = str(row[code_col - 1]).strip() if row[code_col - 1] else ""
                
                if tracking_code and cargo_code and "nan" not in tracking_code.lower():
                    match = re.match(r"^TJ-(\d+)-(\d+)$", cargo_code)
                    if match:
                        day = match.group(1)
                        arrival_date = f"{int(day):02d}.{datetime.now().month:02d}.{datetime.now().year}"
                        extracted_data.append((tracking_code, arrival_date))
            
        elif file_name.endswith(".csv"):
            with open(file_path, encoding='utf-8') as f:
                reader = csv.DictReader(f)
                if "Трек-коды" not in reader.fieldnames or "Код номер" not in reader.fieldnames:
                    await update.reply("❌ Неверный формат CSV! В файле должны быть столбцы ‘Трек-коды’ и ‘Код номер’.")
                    return
                
                for row in reader:
                    tracking_code = row.get("Трек-коды", "").strip()
                    cargo_code = row.get("Код номер", "").strip()
                    
                    if tracking_code and cargo_code and "nan" not in tracking_code.lower():
                        match = re.match(r"^TJ-(\d+)-(\d+)$", cargo_code)
                        if match:
                            day = match.group(1)
                            arrival_date = f"{int(day):02d}.{datetime.now().month:02d}.{datetime.now().year}"
                            extracted_data.append((tracking_code, arrival_date))
        
        # ✅ Delete records from database
        if extracted_data:
            await delete_from_db(extracted_data)
            unique_count = len(extracted_data)
            await update.reply(f"✅ {unique_count} записи удалены.")
            keyboard = [
            ["📦 Обновить список трек-кодов", "📈 Статистика"],
            ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
            ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
            ["📤 Экспорт данных"]
            ]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
            await state.finish()
        else:
            await update.reply("⚠️ В файле не найдено действительных трек-кодов.")
            keyboard = [
            ["📦 Обновить список трек-кодов", "📈 Статистика"],
            ["🔔 Уведомить о прибытии трек-кодов","♻️ Изменение статуса трек-кодов"],
            ["🗑 Удалить завершенные треки", "📝 Добавить трек-код вручную"],
            ["📤 Экспорт данных"]
            ]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.answer("📌 Выберите вариант:", reply_markup=reply_markup)
            await state.finish()

        # ✅ Delete file after processing
        os.remove(file_path)
    
    except Exception as e:
        print(f"❌ Error processing Excel file: {e}")
        await update.reply("❌ Не удалось обработать файл. Проверьте формат и попробуйте еще раз.")
